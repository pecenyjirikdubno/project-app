from flask import (
    Flask,
    render_template,
    request,
    redirect,
    send_file,
    url_for,
    flash,
    send_from_directory,
    jsonify,
)
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager,
    login_user,
    login_required,
    logout_user,
    UserMixin,
    current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from datetime import datetime, date
from zoneinfo import ZoneInfo
import io
import json
import os
import time
import hmac
import hashlib

app = Flask(__name__)
app.config["SECRET_KEY"] = "secret-key-change-this"

# =====================
# DATABASE
# =====================

database_url = os.environ.get("DATABASE_URL")

if not database_url:
    database_url = "sqlite:///database.db"

if database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# =====================
# LOGIN MANAGER
# =====================

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "login"
login_manager.login_message = "Nejdřív se prosím přihlas."

# =====================
# TIMEZONE + DYNAMIC QR
# =====================

APP_TZ = ZoneInfo("Europe/Prague")

QR_PREFIX = "JZQR1"
QR_ROTATION_SECONDS = 30
QR_ALLOWED_SLOT_SKEW = 1
QR_SECRET = os.environ.get("QR_SECRET", "zmenit-v-production-na-dlouhy-tajny-klic")


def now_local():
    return datetime.now(APP_TZ)


def today_local():
    return now_local().date()


def first_day_of_month(year: int, month: int) -> date:
    return date(year, month, 1)


def next_month_first_day(year: int, month: int) -> date:
    if month == 12:
        return date(year + 1, 1, 1)
    return date(year, month + 1, 1)


def time_to_str(value):
    return value.strftime("%H:%M") if value else ""


def datetime_to_str(value):
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def parse_time_hhmm(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%H:%M").time()
    except ValueError:
        return None


def parse_date_yyyy_mm_dd(value: str):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def parse_datetime_value(value: str):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def current_qr_slot(ts=None):
    if ts is None:
        ts = time.time()
    return int(ts // QR_ROTATION_SECONDS)


def build_dynamic_qr_value(slot: int) -> str:
    payload = f"{QR_PREFIX}|{slot}"
    signature = hmac.new(
        QR_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()[:20]
    return f"{QR_PREFIX}|{slot}|{signature}"


def verify_dynamic_qr_value(value: str):
    if not value:
        return False, "Prázdný QR kód."

    parts = value.strip().split("|")
    if len(parts) != 3:
        return False, "Neplatný QR kód."

    prefix, slot_raw, signature = parts

    if prefix != QR_PREFIX:
        return False, "Neplatný QR kód."

    if not slot_raw.isdigit():
        return False, "Neplatný QR kód."

    slot = int(slot_raw)
    now_slot = current_qr_slot()

    if abs(now_slot - slot) > QR_ALLOWED_SLOT_SKEW:
        return False, "QR kód vypršel. Načtěte aktuální kód."

    expected_payload = f"{QR_PREFIX}|{slot}"
    expected_signature = hmac.new(
        QR_SECRET.encode("utf-8"),
        expected_payload.encode("utf-8"),
        hashlib.sha256
    ).hexdigest()[:20]

    if not hmac.compare_digest(signature, expected_signature):
        return False, "Neplatný QR kód."

    return True, None


# =====================
# MODELY
# =====================

class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="user")

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)


class Job(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    closed = db.Column(db.Boolean, default=False, nullable=False)


class JobRow(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    job_id = db.Column(db.Integer, db.ForeignKey("job.id"), nullable=False)

    date = db.Column(db.String(20))
    material_name = db.Column(db.String(200))
    quantity = db.Column(db.Float)
    document_number = db.Column(db.String(100))
    km = db.Column(db.Float)
    travel_time = db.Column(db.Float)
    work_hours = db.Column(db.Float)


class Attendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    work_date = db.Column(db.Date, nullable=False)

    start_time = db.Column(db.Time, nullable=True)
    end_time = db.Column(db.Time, nullable=True)

    created_at = db.Column(db.DateTime, default=lambda: now_local(), nullable=False)
    updated_at = db.Column(db.DateTime, default=lambda: now_local(), nullable=False)

    start_recorded_at = db.Column(db.DateTime, nullable=True)
    end_recorded_at = db.Column(db.DateTime, nullable=True)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# =====================
# INIT DB
# =====================

with app.app_context():
    db.create_all()

    if not User.query.filter_by(username="admin").first():
        admin = User(username="admin", role="admin")
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()


# =====================
# HELPERS
# =====================

def admin_required():
    if current_user.role != "admin":
        flash("Tato akce je dostupná jen pro admina.", "error")
        return False
    return True


def qr_terminal_required():
    if current_user.role not in {"admin", "qr_terminal"}:
        flash("Tato stránka je dostupná jen pro QR terminál.", "error")
        return False
    return True


def user_app_access_required():
    if current_user.role == "qr_terminal":
        flash("Tento účet má přístup pouze k QR terminálu.", "error")
        return False
    return True


def can_user_edit_attendance(record: Attendance) -> bool:
    if current_user.role == "admin":
        return True
    return record.user_id == current_user.id and record.work_date == today_local()


def redirect_after_login():
    if current_user.role == "qr_terminal":
        return redirect(url_for("qr_display"))
    return redirect(url_for("dashboard"))


def export_backup_data():
    users = User.query.order_by(User.id.asc()).all()
    jobs = Job.query.order_by(Job.id.asc()).all()
    job_rows = JobRow.query.order_by(JobRow.id.asc()).all()
    attendance = Attendance.query.order_by(Attendance.id.asc()).all()

    data = {
        "meta": {
            "created_at": now_local().isoformat(),
            "app": "JZ Elektro evidenční systém",
            "format": "internal-json-backup-v1",
        },
        "users": [
            {
                "id": u.id,
                "username": u.username,
                "password_hash": u.password_hash,
                "role": u.role,
            }
            for u in users
        ],
        "jobs": [
            {
                "id": j.id,
                "name": j.name,
                "closed": j.closed,
            }
            for j in jobs
        ],
        "job_rows": [
            {
                "id": r.id,
                "job_id": r.job_id,
                "date": r.date,
                "material_name": r.material_name,
                "quantity": r.quantity,
                "document_number": r.document_number,
                "km": r.km,
                "travel_time": r.travel_time,
                "work_hours": r.work_hours,
            }
            for r in job_rows
        ],
        "attendance": [
            {
                "id": a.id,
                "user_id": a.user_id,
                "work_date": a.work_date.isoformat() if a.work_date else None,
                "start_time": a.start_time.strftime("%H:%M:%S") if a.start_time else None,
                "end_time": a.end_time.strftime("%H:%M:%S") if a.end_time else None,
                "created_at": a.created_at.isoformat() if a.created_at else None,
                "updated_at": a.updated_at.isoformat() if a.updated_at else None,
                "start_recorded_at": a.start_recorded_at.isoformat() if a.start_recorded_at else None,
                "end_recorded_at": a.end_recorded_at.isoformat() if a.end_recorded_at else None,
            }
            for a in attendance
        ],
    }
    return data


def restore_backup_data(data):
    if not isinstance(data, dict):
        raise ValueError("Záloha nemá správný formát.")

    required_keys = {"users", "jobs", "job_rows", "attendance"}
    if not required_keys.issubset(set(data.keys())):
        raise ValueError("Záloha neobsahuje všechny potřebné části.")

    db.session.remove()

    Attendance.query.delete()
    JobRow.query.delete()
    Job.query.delete()
    User.query.delete()
    db.session.commit()

    max_user_id = 0
    max_job_id = 0
    max_job_row_id = 0
    max_attendance_id = 0

    for item in data.get("users", []):
        user = User(
            id=item["id"],
            username=item["username"],
            password_hash=item["password_hash"],
            role=item["role"],
        )
        db.session.add(user)
        max_user_id = max(max_user_id, item["id"])

    for item in data.get("jobs", []):
        job = Job(
            id=item["id"],
            name=item["name"],
            closed=item["closed"],
        )
        db.session.add(job)
        max_job_id = max(max_job_id, item["id"])

    for item in data.get("job_rows", []):
        row = JobRow(
            id=item["id"],
            job_id=item["job_id"],
            date=item.get("date"),
            material_name=item.get("material_name"),
            quantity=item.get("quantity"),
            document_number=item.get("document_number"),
            km=item.get("km"),
            travel_time=item.get("travel_time"),
            work_hours=item.get("work_hours"),
        )
        db.session.add(row)
        max_job_row_id = max(max_job_row_id, item["id"])

    for item in data.get("attendance", []):
        start_time = None
        end_time = None

        if item.get("start_time"):
            start_time = datetime.strptime(item["start_time"], "%H:%M:%S").time()
        if item.get("end_time"):
            end_time = datetime.strptime(item["end_time"], "%H:%M:%S").time()

        attendance = Attendance(
            id=item["id"],
            user_id=item["user_id"],
            work_date=parse_date_yyyy_mm_dd(item.get("work_date")),
            start_time=start_time,
            end_time=end_time,
            created_at=parse_datetime_value(item.get("created_at")) or now_local(),
            updated_at=parse_datetime_value(item.get("updated_at")) or now_local(),
            start_recorded_at=parse_datetime_value(item.get("start_recorded_at")),
            end_recorded_at=parse_datetime_value(item.get("end_recorded_at")),
        )
        db.session.add(attendance)
        max_attendance_id = max(max_attendance_id, item["id"])

    db.session.commit()

    try:
        engine_name = db.engine.url.drivername
        if "postgresql" in engine_name:
            with db.engine.begin() as conn:
                conn.execute(
                    db.text(f"SELECT setval(pg_get_serial_sequence('user', 'id'), {max(max_user_id, 1)}, true);")
                )
                conn.execute(
                    db.text(f"SELECT setval(pg_get_serial_sequence('job', 'id'), {max(max_job_id, 1)}, true);")
                )
                conn.execute(
                    db.text(f"SELECT setval(pg_get_serial_sequence('job_row', 'id'), {max(max_job_row_id, 1)}, true);")
                )
                conn.execute(
                    db.text(f"SELECT setval(pg_get_serial_sequence('attendance', 'id'), {max(max_attendance_id, 1)}, true);")
                )
    except Exception:
        pass


# =====================
# PWA
# =====================

@app.route("/manifest.json")
def manifest():
    return send_from_directory(".", "manifest.json", mimetype="application/manifest+json")


@app.route("/service-worker.js")
def service_worker():
    return send_from_directory(".", "service-worker.js", mimetype="application/javascript")


# =====================
# AUTH
# =====================

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect_after_login()

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user)
            return redirect_after_login()

        flash("Neplatné uživatelské jméno nebo heslo.", "error")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# =====================
# MAIN DASHBOARD
# =====================

@app.route("/")
@login_required
def dashboard():
    if current_user.role == "qr_terminal":
        return redirect(url_for("qr_display"))
    return render_template("dashboard.html")


@app.route("/qr-display")
@login_required
def qr_display():
    if not qr_terminal_required():
        return redirect(url_for("dashboard"))

    slot = current_qr_slot()
    qr_value = build_dynamic_qr_value(slot)

    seconds_into_slot = int(time.time()) % QR_ROTATION_SECONDS
    refresh_after = max(3, QR_ROTATION_SECONDS - seconds_into_slot + 1)

    return render_template(
        "qr_display.html",
        qr_value=qr_value,
        refresh_after=refresh_after,
        refresh_ms=refresh_after * 1000,
        qr_rotation_seconds=QR_ROTATION_SECONDS,
    )


# =====================
# MATERIÁL
# =====================

@app.route("/materials")
@login_required
def materials():
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    jobs = Job.query.order_by(Job.id.desc()).all()

    for job in jobs:
        job.rows = JobRow.query.filter_by(job_id=job.id).all()
        job.total_quantity = sum((r.quantity or 0) for r in job.rows)
        job.total_km = sum((r.km or 0) for r in job.rows)
        job.total_travel_time = sum((r.travel_time or 0) for r in job.rows)
        job.total_work_hours = sum((r.work_hours or 0) for r in job.rows)

    return render_template("materials.html", jobs=jobs)


@app.route("/create_job", methods=["POST"])
@login_required
def create_job():
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    name = request.form.get("name", "").strip()

    if not name:
        flash("Zadej název zakázky.", "error")
        return redirect(url_for("materials"))

    new_job = Job(name=name)
    db.session.add(new_job)
    db.session.commit()

    return redirect(url_for("materials"))


@app.route("/add_row/<int:job_id>", methods=["POST"])
@login_required
def add_row(job_id):
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    job = Job.query.get(job_id)

    if not job or job.closed:
        return redirect(url_for("materials"))

    new_row = JobRow(
        job_id=job_id,
        date="",
        material_name="",
        quantity=0,
        document_number="",
        km=0,
        travel_time=0,
        work_hours=0,
    )

    db.session.add(new_row)
    db.session.commit()

    return redirect(url_for("materials"))


@app.route("/save/<int:job_id>", methods=["POST"])
@login_required
def save(job_id):
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    job = Job.query.get(job_id)

    if not job or job.closed:
        return redirect(url_for("materials"))

    rows = JobRow.query.filter_by(job_id=job_id).all()

    for row in rows:
        row.date = request.form.get(f"date_{row.id}")
        row.material_name = request.form.get(f"material_name_{row.id}")
        row.quantity = float(request.form.get(f"quantity_{row.id}") or 0)
        row.document_number = request.form.get(f"document_number_{row.id}")
        row.km = float(request.form.get(f"km_{row.id}") or 0)
        row.travel_time = float(request.form.get(f"travel_time_{row.id}") or 0)
        row.work_hours = float(request.form.get(f"work_hours_{row.id}") or 0)

    db.session.commit()
    flash("Zakázka uložena.", "success")
    return redirect(url_for("materials"))


@app.route("/close/<int:job_id>")
@login_required
def close_job(job_id):
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    job = Job.query.get(job_id)

    if job and current_user.role == "admin":
        job.closed = True
        db.session.commit()

    return redirect(url_for("materials"))


@app.route("/export/<int:job_id>")
@login_required
def export(job_id):
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    rows = JobRow.query.filter_by(job_id=job_id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Materiál"

    ws.append([
        "Datum",
        "Materiál",
        "Množství",
        "Číslo dokladu",
        "Km",
        "Čas na cestě",
        "Odpracované hodiny",
    ])

    for r in rows:
        ws.append([
            r.date,
            r.material_name,
            r.quantity,
            r.document_number,
            r.km,
            r.travel_time,
            r.work_hours,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"zakazka_{job_id}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# =====================
# DOCHÁZKA
# =====================

@app.route("/attendance")
@login_required
def attendance():
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    selected_month = request.args.get("month")

    if selected_month:
        try:
            year, month = map(int, selected_month.split("-"))
        except ValueError:
            current = today_local()
            year, month = current.year, current.month
    else:
        current = today_local()
        year, month = current.year, current.month

    month_start = first_day_of_month(year, month)
    month_end = next_month_first_day(year, month)

    if current_user.role == "admin":
        records = (
            Attendance.query.filter(
                Attendance.work_date >= month_start,
                Attendance.work_date < month_end,
            )
            .order_by(Attendance.work_date.desc(), Attendance.id.desc())
            .all()
        )
    else:
        records = (
            Attendance.query.filter(
                Attendance.user_id == current_user.id,
                Attendance.work_date >= month_start,
                Attendance.work_date < month_end,
            )
            .order_by(Attendance.work_date.desc(), Attendance.id.desc())
            .all()
        )

    user_map = {}
    if current_user.role == "admin":
        user_ids = {r.user_id for r in records}
        if user_ids:
            users = User.query.filter(User.id.in_(user_ids)).all()
            user_map = {u.id: u.username for u in users}

    today_record = Attendance.query.filter_by(
        user_id=current_user.id,
        work_date=today_local()
    ).first()

    return render_template(
        "attendance.html",
        records=records,
        today=today_local().isoformat(),
        selected_month=f"{year:04d}-{month:02d}",
        user_map=user_map,
        can_user_edit_attendance=can_user_edit_attendance,
        time_to_str=time_to_str,
        today_record=today_record,
    )


@app.route("/attendance/create_day", methods=["POST"])
@login_required
def create_attendance_day():
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    work_date_raw = request.form.get("work_date", "").strip()

    if not work_date_raw:
        flash("Vyber datum.", "error")
        return redirect(url_for("attendance"))

    work_date = parse_date_yyyy_mm_dd(work_date_raw)
    if not work_date:
        flash("Neplatné datum.", "error")
        return redirect(url_for("attendance"))

    existing = Attendance.query.filter_by(
        user_id=current_user.id,
        work_date=work_date
    ).first()

    if existing:
        flash("Záznam pro toto datum už existuje.", "error")
        return redirect(url_for("attendance"))

    record = Attendance(
        user_id=current_user.id,
        work_date=work_date,
        created_at=now_local(),
        updated_at=now_local(),
    )
    db.session.add(record)
    db.session.commit()

    flash("Den docházky byl vytvořen.", "success")
    return redirect(url_for("attendance"))


@app.route("/attendance/scan-qr", methods=["POST"])
@login_required
def attendance_scan_qr():
    if current_user.role == "qr_terminal":
        return jsonify({
            "success": False,
            "message": "Terminálový účet nemůže zapisovat docházku.",
        }), 403

    payload = request.get_json(silent=True) or {}
    qr_value = (payload.get("qr_value") or "").strip()

    is_valid, error_message = verify_dynamic_qr_value(qr_value)
    if not is_valid:
        return jsonify({
            "success": False,
            "message": error_message,
        }), 400

    work_date = today_local()
    current_dt = now_local()
    current_tm = current_dt.time().replace(second=0, microsecond=0)

    record = Attendance.query.filter_by(
        user_id=current_user.id,
        work_date=work_date
    ).first()

    if not record:
        record = Attendance(
            user_id=current_user.id,
            work_date=work_date,
            start_time=current_tm,
            created_at=current_dt,
            updated_at=current_dt,
            start_recorded_at=current_dt,
        )
        db.session.add(record)
        db.session.commit()

        return jsonify({
            "success": True,
            "action": "start",
            "message": f"Příchod zapsán v {record.start_time.strftime('%H:%M')}.",
        })

    if current_user.role != "admin" and record.work_date != today_local():
        return jsonify({
            "success": False,
            "message": "Tento záznam už nelze upravit.",
        }), 403

    if record.start_time is None:
        record.start_time = current_tm
        record.start_recorded_at = current_dt
        record.updated_at = current_dt
        db.session.commit()

        return jsonify({
            "success": True,
            "action": "start",
            "message": f"Příchod zapsán v {record.start_time.strftime('%H:%M')}.",
        })

    if record.end_time is None:
        record.end_time = current_tm
        record.end_recorded_at = current_dt
        record.updated_at = current_dt
        db.session.commit()

        return jsonify({
            "success": True,
            "action": "end",
            "message": f"Odchod zapsán v {record.end_time.strftime('%H:%M')}.",
        })

    return jsonify({
        "success": False,
        "message": "Docházka pro dnešek je již uzavřena.",
    }), 400


@app.route("/attendance/user_update/<int:record_id>", methods=["POST"])
@login_required
def attendance_user_update(record_id):
    if not user_app_access_required():
        return redirect(url_for("qr_display"))

    record = Attendance.query.get_or_404(record_id)

    if not can_user_edit_attendance(record):
        flash("Tento záznam už nemůžeš upravovat.", "error")
        return redirect(url_for("attendance"))

    start_time_raw = request.form.get("start_time", "").strip()
    end_time_raw = request.form.get("end_time", "").strip()

    new_start = parse_time_hhmm(start_time_raw)
    new_end = parse_time_hhmm(end_time_raw)

    if start_time_raw and new_start is None:
        flash("Neplatný čas nástupu.", "error")
        return redirect(url_for("attendance"))

    if end_time_raw and new_end is None:
        flash("Neplatný čas ukončení.", "error")
        return redirect(url_for("attendance"))

    if record.start_time != new_start and new_start is not None:
        record.start_recorded_at = now_local()

    if record.end_time != new_end and new_end is not None:
        record.end_recorded_at = now_local()

    record.start_time = new_start
    record.end_time = new_end
    record.updated_at = now_local()

    db.session.commit()
    flash("Docházka uložena.", "success")
    return redirect(url_for("attendance"))


@app.route("/attendance/admin_update/<int:record_id>", methods=["POST"])
@login_required
def attendance_admin_update(record_id):
    if not admin_required():
        return redirect(url_for("attendance"))

    record = Attendance.query.get_or_404(record_id)

    work_date_raw = request.form.get("work_date", "").strip()
    start_time_raw = request.form.get("start_time", "").strip()
    end_time_raw = request.form.get("end_time", "").strip()

    work_date = parse_date_yyyy_mm_dd(work_date_raw)
    if not work_date:
        flash("Neplatné datum.", "error")
        return redirect(url_for("attendance"))

    new_start = parse_time_hhmm(start_time_raw)
    new_end = parse_time_hhmm(end_time_raw)

    if start_time_raw and new_start is None:
        flash("Neplatný čas nástupu.", "error")
        return redirect(url_for("attendance"))

    if end_time_raw and new_end is None:
        flash("Neplatný čas ukončení.", "error")
        return redirect(url_for("attendance"))

    if record.start_time != new_start and new_start is not None:
        record.start_recorded_at = now_local()

    if record.end_time != new_end and new_end is not None:
        record.end_recorded_at = now_local()

    record.work_date = work_date
    record.start_time = new_start
    record.end_time = new_end
    record.updated_at = now_local()

    db.session.commit()
    flash("Docházka upravena.", "success")
    return redirect(url_for("attendance"))


@app.route("/attendance/export/all/excel")
@login_required
def attendance_export_all_excel():
    if not admin_required():
        return redirect(url_for("attendance"))

    records = Attendance.query.order_by(Attendance.work_date.desc(), Attendance.id.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Docházka"

    ws.append([
        "Uživatel",
        "Datum docházky",
        "Nástup",
        "Ukončení",
        "Vytvořeno",
        "Naposledy upraveno",
        "Zápis nástupu proveden",
        "Zápis ukončení proveden",
    ])

    user_ids = {r.user_id for r in records}
    users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {u.id: u.username for u in users}

    for r in records:
        ws.append([
            user_map.get(r.user_id, ""),
            r.work_date.isoformat() if r.work_date else "",
            time_to_str(r.start_time),
            time_to_str(r.end_time),
            datetime_to_str(r.created_at),
            datetime_to_str(r.updated_at),
            datetime_to_str(r.start_recorded_at),
            datetime_to_str(r.end_recorded_at),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="dochazka_vse.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/attendance/export/monthly/excel")
@login_required
def attendance_export_monthly_excel():
    if not admin_required():
        return redirect(url_for("attendance"))

    month_raw = request.args.get("month")
    if not month_raw:
        month_raw = today_local().strftime("%Y-%m")

    year, month = map(int, month_raw.split("-"))
    month_start = first_day_of_month(year, month)
    month_end = next_month_first_day(year, month)

    records = (
        Attendance.query.filter(
            Attendance.work_date >= month_start,
            Attendance.work_date < month_end,
        )
        .order_by(Attendance.work_date.asc(), Attendance.id.asc())
        .all()
    )

    user_ids = {r.user_id for r in records}
    users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {u.id: u.username for u in users}

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    ws.append([
        "Uživatel",
        "Datum docházky",
        "Nástup",
        "Ukončení",
        "Vytvořeno",
        "Naposledy upraveno",
        "Zápis nástupu proveden",
        "Zápis ukončení proveden",
    ])

    for r in records:
        ws.append([
            user_map.get(r.user_id, ""),
            r.work_date.isoformat() if r.work_date else "",
            time_to_str(r.start_time),
            time_to_str(r.end_time),
            datetime_to_str(r.created_at),
            datetime_to_str(r.updated_at),
            datetime_to_str(r.start_recorded_at),
            datetime_to_str(r.end_recorded_at),
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"dochazka_{year}_{month:02d}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/attendance/export/monthly/pdf")
@login_required
def attendance_export_monthly_pdf():
    if not admin_required():
        return redirect(url_for("attendance"))

    month_raw = request.args.get("month")
    if not month_raw:
        month_raw = today_local().strftime("%Y-%m")

    year, month = map(int, month_raw.split("-"))
    month_start = first_day_of_month(year, month)
    month_end = next_month_first_day(year, month)

    records = (
        Attendance.query.filter(
            Attendance.work_date >= month_start,
            Attendance.work_date < month_end,
        )
        .order_by(Attendance.work_date.asc(), Attendance.id.asc())
        .all()
    )

    user_ids = {r.user_id for r in records}
    users = User.query.filter(User.id.in_(user_ids)).all() if user_ids else []
    user_map = {u.id: u.username for u in users}

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph(f"Docházka za měsíc {year}-{month:02d}", styles["Title"]))
    elements.append(Spacer(1, 12))

    table_data = [[
        "Uživatel",
        "Datum",
        "Nástup",
        "Ukončení",
        "Vytvořeno",
        "Upraveno",
        "Zápis nástupu",
        "Zápis ukončení",
    ]]

    for r in records:
        table_data.append([
            user_map.get(r.user_id, ""),
            r.work_date.isoformat() if r.work_date else "",
            time_to_str(r.start_time),
            time_to_str(r.end_time),
            datetime_to_str(r.created_at),
            datetime_to_str(r.updated_at),
            datetime_to_str(r.start_recorded_at),
            datetime_to_str(r.end_recorded_at),
        ])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"dochazka_{year}_{month:02d}.pdf",
        mimetype="application/pdf",
    )


# =====================
# ADMIN INTERNAL BACKUP / RESTORE
# =====================

@app.route("/admin/db-backup")
@login_required
def admin_db_backup():
    if not admin_required():
        return redirect(url_for("dashboard"))

    backup_data = export_backup_data()

    output = io.BytesIO()
    output.write(json.dumps(backup_data, ensure_ascii=False, indent=2).encode("utf-8"))
    output.seek(0)

    filename = f"backup_{now_local().strftime('%Y-%m-%d_%H-%M-%S')}.json"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/json",
    )


@app.route("/admin/db-restore", methods=["POST"])
@login_required
def admin_db_restore():
    if not admin_required():
        return redirect(url_for("dashboard"))

    uploaded_file = request.files.get("backup_file")

    if not uploaded_file or uploaded_file.filename == "":
        flash("Vyber JSON soubor pro obnovu.", "error")
        return redirect(url_for("dashboard"))

    if not uploaded_file.filename.lower().endswith(".json"):
        flash("Obnovit lze jen ze souboru .json", "error")
        return redirect(url_for("dashboard"))

    try:
        content = uploaded_file.read().decode("utf-8")
        data = json.loads(content)
        restore_backup_data(data)
        flash("Databáze byla úspěšně obnovena z interní zálohy.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Obnova databáze selhala: {str(e)}", "error")

    return redirect(url_for("dashboard"))


# =====================
# USER MANAGEMENT
# =====================

@app.route("/users")
@login_required
def users():
    if not admin_required():
        return redirect(url_for("dashboard"))

    all_users = User.query.order_by(User.id.asc()).all()
    return render_template("users.html", users=all_users)


@app.route("/add_user", methods=["POST"])
@login_required
def add_user():
    if not admin_required():
        return redirect(url_for("dashboard"))

    username = request.form.get("username", "").strip()
    password = request.form.get("password", "")
    role = request.form.get("role", "user")

    if not username or not password:
        flash("Vyplň uživatelské jméno i heslo.", "error")
        return redirect(url_for("users"))

    if role not in {"admin", "user", "qr_terminal"}:
        role = "user"

    if User.query.filter_by(username=username).first():
        flash("Uživatel už existuje.", "error")
        return redirect(url_for("users"))

    new_user = User(username=username, role=role)
    new_user.set_password(password)

    db.session.add(new_user)
    db.session.commit()

    flash("Uživatel byl vytvořen.", "success")
    return redirect(url_for("users"))


@app.route("/delete_user/<int:user_id>")
@login_required
def delete_user(user_id):
    if not admin_required():
        return redirect(url_for("dashboard"))

    user = User.query.get_or_404(user_id)

    if user.username == "admin":
        flash("Hlavního admina nelze smazat.", "error")
        return redirect(url_for("users"))

    db.session.delete(user)
    db.session.commit()

    flash("Uživatel byl smazán.", "success")
    return redirect(url_for("users"))


# =====================
# RUN
# =====================

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
